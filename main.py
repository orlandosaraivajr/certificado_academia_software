from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from datetime import datetime
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
import os


def create_certificate(nome_aluno, curso, carga_horaria, responsavel, cargo_responsavel, data_evento, data_evento_completo):
    os.makedirs('./todos', exist_ok=True)

    # Tamanho personalizado (4959 x 7016 pontos)
    page_size = (4959, 7016)

    certificado_base = "certificado_template.pdf"
    certificado_temp = f"./todos/temp_certificado_{nome_aluno.replace(' ', '_')}.pdf"
    certificado_final = f"./todos/{nome_aluno.replace(' ', '_')}.pdf"

    c = canvas.Canvas(certificado_temp, pagesize=page_size)
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 50)
    c.drawCentredString(550, 1255, f"Certificamos que")

    c.setFont("Times-Roman", 40)
    c.drawString(50, 1100, f" {nome_aluno}")

    c.drawString(50, 1100 - (1*70), f" participou e concluiu com êxito o treinamento")

    c.drawString(50, 1100 - (2*70), f" {curso}")
    c.drawString(50, 1100 - (3*70), f" realizado no dia {data_evento} com duração de {carga_horaria} horas")


    c.drawString(600, 700, f" Rio Claro,  {data_evento_completo} ")
    
    c.setFont("Times-Roman", 30)
    c.drawString(250, 500, 30 * "_")
    c.drawString(250, 455, f"{responsavel}")
    c.drawString(250, 405, f"{cargo_responsavel}")

    c.setFont("Times-Roman", 60)
    c.drawCentredString(page_size[0] / 2, 3800, f"Concluiu o curso de {curso}")
    c.drawCentredString(page_size[0] / 2, 3600, f"Carga horária: {int(carga_horaria)} horas")

    data_emissao = datetime.now().strftime("%d/%m/%Y")
    c.setFillColor(colors.white)
    c.setFont("Times-Roman", 20)
    c.drawString(900, 30, f"Data de emissão: {data_emissao}")
    

    c.save()

    # Mescla com o template base
    background = PdfReader(certificado_base)
    overlay = PdfReader(certificado_temp)
    output = PdfWriter()

    background_page = background.pages[0]
    overlay_page = overlay.pages[0]
    background_page.merge_page(overlay_page)
    output.add_page(background_page)

    with open(certificado_final, "wb") as out_file:
        output.write(out_file)

    os.remove(certificado_temp)

    print(f"✅ Certificado gerado: {certificado_final}")


# Dados Evento
planilha_alunos = load_workbook('./alunos.xlsx')
evento_sheet = planilha_alunos['Evento']
curso = evento_sheet['B1'].value
carga_horaria = evento_sheet['B2'].value
responsavel = evento_sheet['B3'].value
cargo_responsavel = evento_sheet['B4'].value
data_evento = '28/06/2025'
data_evento_completo = '28 de junho de 2025'
pagina_certificados = planilha_alunos['certificados']

# Gera certificados
for linha in pagina_certificados.iter_rows(min_row=2, values_only=True):
    nome_aluno, origem = linha
    create_certificate(nome_aluno, curso, carga_horaria, responsavel, cargo_responsavel, data_evento, data_evento_completo)
